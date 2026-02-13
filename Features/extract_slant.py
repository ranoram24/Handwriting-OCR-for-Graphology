import cv2
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Tuple
import argparse
from tqdm import tqdm
from openpyxl import load_workbook

def measure_slant_by_shear(binary_img: np.ndarray) -> Tuple[float, float]:
    # Calculates the slant by applying affine shear transformations to the image.
    # Logic: We shear the image at different angles. The angle that maximizes the 
    # variance of the vertical projection is the one that makes the text most upright.
    
    img_height, img_width = binary_img.shape
    
    # Test angles from -30 to +30 degrees
    angles_to_test = np.linspace(-30, 30, 61)
    max_variance = 0
    optimal_angle = 0

    for angle in angles_to_test:
        angle_rad = np.radians(angle)
        shear_factor = np.tan(angle_rad)
        
        # Define the Affine Transformation Matrix for shearing
        M = np.array([[1, shear_factor, 0],
                      [0, 1, 0]], dtype=np.float32)
        
        # Apply the shear transformation
        sheared = cv2.warpAffine(binary_img, M, (img_width + abs(int(shear_factor * img_height)), img_height),
                                flags=cv2.INTER_LINEAR, borderValue=255)
        
        # Calculate vertical projection (sum of black pixels per column)
        projection = np.sum(sheared == 0, axis=0)
        
        # Calculate variance - higher variance means sharper peaks/valleys, indicating upright text
        variance = np.var(projection)

        if variance > max_variance:
            max_variance = variance
            optimal_angle = angle

    # Normalize the result to 0.0 - 1.0 range
    # -30 deg = 0.0, 0 deg = 0.5, +30 deg = 1.0
    slant = (optimal_angle + 30) / 60
    slant = np.clip(slant, 0.0, 1.0)
    
    return slant, optimal_angle

def measure_slant_by_moments(binary_img: np.ndarray) -> float:
    # Calculates slant by finding contours of individual letters and fitting ellipses.
    # The average angle of these ellipses represents the handwriting slant.
    
    # Invert image to find contours (OpenCV expects white object on black background)
    inverted = 255 - binary_img
    contours, _ = cv2.findContours(inverted, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

    if not contours:
        return 0.5

    angles = []
    img_area = binary_img.shape[0] * binary_img.shape[1]

    for contour in contours:
        area = cv2.contourArea(contour)
        
        # Filter noise: Ignore contours that are too small or too large
        if area < img_area * 0.0001 or area > img_area * 0.15:
            continue
        if len(contour) < 5:
            continue

        try:
            # Fit an ellipse to the contour to determine orientation
            (x, y), (MA, ma), angle = cv2.fitEllipse(contour)
            
            # Adjust angle to be relative to vertical axis
            if angle > 135:
                angle = angle - 180
            elif angle > 45:
                angle = 90 - angle
            angles.append(angle)
        except:
            continue

    if not angles:
        return 0.5

    # Remove statistical outliers using Interquartile Range (IQR) to get a robust mean
    angles = np.array(angles)
    q1, q3 = np.percentile(angles, [25, 75])
    iqr = q3 - q1
    if iqr > 0:
        mask = (angles >= q1 - 1.5*iqr) & (angles <= q3 + 1.5*iqr)
        angles = angles[mask]

    if len(angles) == 0:
        return 0.5

    mean_angle = np.mean(angles)
    
    # Normalize to 0.0 - 1.0 range based on -20 to +20 degrees limits
    mean_angle = np.clip(mean_angle, -20, 20)
    slant = (mean_angle + 20) / 40
    return slant

def extract_slant(image_path: str) -> Tuple[float, dict]:
    # Main function to process a single image.
    # Combines Shear method (global) and Moments method (local) for better accuracy.
    
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)
    if img is None:
        return 0.5, {'error': 'Failed to read'}

    # Ensure binary image (Thresholding) if not already
    if len(np.unique(img)) > 2:
        _, img = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY)

    # Run both algorithms
    slant_shear, optimal_angle = measure_slant_by_shear(img)
    slant_moments = measure_slant_by_moments(img)

    # Weighted Average: Giving more weight (0.7) to the Shear method as it is generally more robust for lines
    slant = 0.7 * slant_shear + 0.3 * slant_moments
    slant = np.clip(slant, 0.0, 1.0)

    debug_info = {
        'slant_shear': slant_shear,
        'slant_moments': slant_moments,
        'optimal_angle': optimal_angle
    }
    return slant, debug_info

def process_directory(input_dir: str, output_file: str) -> pd.DataFrame:
    # Iterates over all images in the directory, calculates slant, and saves to Excel.
    
    image_files = []
    # Collect all common image formats
    for ext in ['.png', '.jpg', '.jpeg', '.tif', '.tiff']:
        image_files.extend(Path(input_dir).glob(f'*{ext}'))
        image_files.extend(Path(input_dir).glob(f'*{ext.upper()}'))

    image_files = sorted(set(image_files))
    total = len(image_files)

    print(f"Found {total} images")
    print(f"Output: {output_file}")
    
    results = []
    # Process images with a progress bar
    for idx, img_path in enumerate(tqdm(image_files, desc="Processing"), 1):
        slant, debug = extract_slant(str(img_path))
        results.append({
            'Image Number': idx,
            'File Name': img_path.name,
            'Slant': round(slant, 3)
        })

    # Save initial data to Excel
    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Post-processing: Add Excel formula to extract image index from filename if needed
    wb = load_workbook(output_file)
    ws = wb.active
    for row_num in range(2, len(results) + 2):
        ws[f'A{row_num}'] = f'=VALUE(MID(B{row_num}, FIND("(",B{row_num})+1, FIND(")",B{row_num})-FIND("(",B{row_num})-1))'
    wb.save(output_file)

    print(f"\nStatistics:")
    print(f"  Mean: {df['Slant'].mean():.3f}")
    print(f"  Median: {df['Slant'].median():.3f}")
    print(f"  Std Dev: {df['Slant'].std():.3f}")
    print(f"  Min: {df['Slant'].min():.3f}")
    print(f"  Max: {df['Slant'].max():.3f}")
    return df

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('input_dir', type=str, help='Path to the folder containing images')
    parser.add_argument('output_file', type=str, help='Path for the output Excel file')
    args = parser.parse_args()
    
    process_directory(args.input_dir, args.output_file)

if __name__ == '__main__':
    main()