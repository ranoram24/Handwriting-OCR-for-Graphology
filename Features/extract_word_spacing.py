import cv2
import numpy as np
import pandas as pd
from pathlib import Path
from typing import Tuple, List
import argparse
from tqdm import tqdm

def find_words_morphological(binary_img: np.ndarray) -> List[Tuple[int, int, int, int]]:
    """
    Find word bounding boxes using morphological operations.
    Returns list of (x, y, w, h) for each word.
    """
    # Invert (text should be white on black for morphology)
    inverted = 255 - binary_img

    # Estimate letter width first to adaptively size kernel
    letter_width = estimate_letter_width_from_image(binary_img)

    # Horizontal dilation to connect letters within words
    # Kernel width based on estimated letter width
    # Use ~0.7x letter width to connect letters but not words
    kernel_width = max(int(letter_width * 0.7), 5)
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (kernel_width, 1))

    # Dilate horizontally
    dilated = cv2.dilate(inverted, kernel, iterations=1)

    # Find connected components
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(
        dilated, connectivity=8
    )

    words = []
    img_area = binary_img.shape[0] * binary_img.shape[1]

    # Skip label 0 (background)
    for i in range(1, num_labels):
        x = stats[i, cv2.CC_STAT_LEFT]
        y = stats[i, cv2.CC_STAT_TOP]
        w = stats[i, cv2.CC_STAT_WIDTH]
        h = stats[i, cv2.CC_STAT_HEIGHT]
        area = stats[i, cv2.CC_STAT_AREA]

        # Filter noise and too-large components
        if area < 20:  # Too small
            continue
        if area > img_area * 0.9:  # Too large
            continue
        if w < 5 or h < 3:  # Too thin
            continue

        words.append((x, y, w, h))

    # Sort by x position
    words.sort(key=lambda b: b[0])

    return words

def estimate_letter_width_from_image(binary_img: np.ndarray) -> float:
    """
    Estimate typical letter width by analyzing the image directly.
    """
    # Find connected components (letters)
    inverted = 255 - binary_img
    num_labels, labels, stats, centroids = cv2.connectedComponentsWithStats(
        inverted, connectivity=8
    )

    widths = []
    for i in range(1, num_labels):
        w = stats[i, cv2.CC_STAT_WIDTH]
        h = stats[i, cv2.CC_STAT_HEIGHT]
        area = stats[i, cv2.CC_STAT_AREA]

        # Filter to get letter-sized components
        if 3 <= w <= 50 and 3 <= h <= 50 and 10 <= area <= 1000:
            widths.append(w)

    if len(widths) == 0:
        return 15.0

    # Use median of letter widths
    return np.median(widths)

def measure_word_spacing(binary_img: np.ndarray) -> Tuple[float, dict]:
    """
    Measure word spacing using morphological word detection.

    Returns:
        Tuple of (spacing_value, debug_info)
    """
    # Ensure text is black (0) and background is white (255)
    if np.mean(binary_img) < 127:
        binary_img = 255 - binary_img

    # Find words using morphology
    words = find_words_morphological(binary_img)

    if len(words) == 0:
        return 0.5, {'note': 'No words found'}

    # Single word
    if len(words) == 1:
        return 0.5, {
            'note': 'Single word',
            'num_words': 1
        }

    # Calculate gaps between words
    gaps = []
    for i in range(len(words) - 1):
        x1, y1, w1, h1 = words[i]
        x2, y2, w2, h2 = words[i + 1]

        gap = x2 - (x1 + w1)
        gaps.append(gap)

    # Filter to positive gaps only
    positive_gaps = [g for g in gaps if g > 0]

    if len(positive_gaps) == 0:
        # Overlapping words - very cramped
        return 0.15, {
            'note': 'Overlapping words - very cramped',
            'num_words': len(words)
        }

    # Calculate average gap
    avg_gap = np.mean(positive_gaps)

    # Estimate letter width
    letter_width = estimate_letter_width_from_image(binary_img)

    # Calculate ratio
    ratio = avg_gap / letter_width if letter_width > 0 else 0.5

    # Map ratio to spacing
    # ratio ~ 0.2 -> spacing ~ 0.0-0.2 (cramped)
    # ratio ~ 1.0 -> spacing ~ 0.5 (normal)
    # ratio ~ 2.5+ -> spacing ~ 1.0 (wide)

    if ratio <= 0.2:
        spacing = ratio * 0.5  # 0-0.2 ratio -> 0-0.1 spacing
    elif ratio <= 1.0:
        spacing = 0.1 + (ratio - 0.2) * 0.4 / 0.8  # 0.2-1.0 ratio -> 0.1-0.5 spacing
    else:
        spacing = 0.5 + min((ratio - 1.0) / 1.5, 0.5)  # 1.0+ ratio -> 0.5-1.0 spacing

    spacing = np.clip(spacing, 0.0, 1.0)

    # Avoid exact endpoints for multi-word
    if spacing < 0.02:
        spacing = 0.01
    elif spacing > 0.98:
        spacing = 0.99

    debug_info = {
        'num_words': len(words),
        'num_gaps': len(gaps),
        'num_positive_gaps': len(positive_gaps),
        'avg_gap': round(avg_gap, 2) if len(positive_gaps) > 0 else 0,
        'letter_width': round(letter_width, 2),
        'ratio': round(ratio, 3)
    }

    return spacing, debug_info

def extract_word_spacing(image_path: str) -> Tuple[float, dict]:
    """
    Extract word spacing feature from image.
    """
    img = cv2.imread(image_path, cv2.IMREAD_GRAYSCALE)

    if img is None:
        return 0.5, {'error': 'Failed to read image'}

    # Ensure binary
    if len(np.unique(img)) > 2:
        _, img = cv2.threshold(img, 127, 255, cv2.THRESH_BINARY)

    spacing, debug = measure_word_spacing(img)

    return spacing, debug

def process_directory(input_dir: str, output_file: str) -> pd.DataFrame:
    """
    Process all images in directory and create Excel file.
    """
    # Get all images
    image_files = []
    for ext in ['.png', '.jpg', '.jpeg', '.tif', '.tiff']:
        image_files.extend(Path(input_dir).glob(f'*{ext}'))
        image_files.extend(Path(input_dir).glob(f'*{ext.upper()}'))

    image_files = sorted(set(image_files))
    total = len(image_files)

    print(f"Found {total} images")
    print(f"Output: {output_file}")
    print("-" * 60)

    results = []

    for idx, img_path in enumerate(tqdm(image_files, desc="Processing"), 1):
        spacing, debug = extract_word_spacing(str(img_path))

        results.append({
            'Image Number': idx,
            'File Name': img_path.name,
            'Word_Spacing': round(spacing, 3)
        })

    df = pd.DataFrame(results)
    df.to_excel(output_file, index=False, engine='openpyxl')

    # Add formula to Image Number column
    from openpyxl import load_workbook
    wb = load_workbook(output_file)
    ws = wb.active

    # Replace Image Number values with formula (starting from row 2)
    for row_num in range(2, len(results) + 2):
        ws[f'A{row_num}'] = f'=VALUE(MID(B{row_num}, FIND("(",B{row_num})+1, FIND(")",B{row_num})-FIND("(",B{row_num})-1))'

    wb.save(output_file)

    print("-" * 60)
    print("Complete!")
    print(f"\nStatistics:")
    print(f"  Mean: {df['Word_Spacing'].mean():.3f}")
    print(f"  Median: {df['Word_Spacing'].median():.3f}")
    print(f"  Std Dev: {df['Word_Spacing'].std():.3f}")
    print(f"  Min: {df['Word_Spacing'].min():.3f}")
    print(f"  Max: {df['Word_Spacing'].max():.3f}")

    # Show distribution
    spacing_05 = len(df[df['Word_Spacing'] == 0.5])
    print(f"  Images with spacing = 0.5 (single word): {spacing_05} ({spacing_05/total*100:.1f}%)")

    return df

def main():
    parser = argparse.ArgumentParser(description='Extract Word Spacing feature')
    parser.add_argument('input_dir', type=str, help='Directory with normalized images')
    parser.add_argument('output_file', type=str, help='Output Excel file path')
    args = parser.parse_args()

    process_directory(args.input_dir, args.output_file)

if __name__ == '__main__':
    main()